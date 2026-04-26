import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        return [[cell for cell in row] for row in reader]


def write_rows_preserve_dimensions(worksheet, rows: list[list[str]]) -> None:
    max_row_existing = worksheet.max_row or 0
    max_col_existing = worksheet.max_column or 0
    max_row_new = len(rows)
    max_col_new = max((len(r) for r in rows), default=0)

    max_row = max(max_row_existing, max_row_new)
    max_col = max(max_col_existing, max_col_new)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            worksheet.cell(row=r, column=c).value = None

    for r_index, row in enumerate(rows, start=1):
        for c_index, value in enumerate(row, start=1):
            worksheet.cell(row=r_index, column=c_index).value = value


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync battlelabs-agent-product-dashboard.xlsx tabs from CSV sources (for deterministic Google Sheets imports).",
    )
    parser.add_argument(
        "--xlsx",
        default=str(Path("ops/agent-native-revenue/battlelabs-agent-product-dashboard.xlsx")),
        help="Path to the dashboard XLSX to update in-place.",
    )
    parser.add_argument(
        "--product-metrics",
        default=str(Path("ops/agent-native-revenue/battlelabs-agent-product-dashboard.csv")),
        help="CSV source for the Product Metrics tab.",
    )
    parser.add_argument(
        "--qa-checklist",
        default=str(Path("ops/agent-native-revenue/dashboard-tabs/QA Checklist.csv")),
        help="CSV source for the QA Checklist tab.",
    )
    parser.add_argument(
        "--task-queue",
        default=str(Path("ops/agent-native-revenue/portfolio-task-queue.csv")),
        help="CSV source for the Task Queue tab.",
    )
    parser.add_argument(
        "--linear-import",
        default=str(Path("ops/agent-native-revenue/linear-import.csv")),
        help="CSV source for the Linear Import tab.",
    )
    parser.add_argument(
        "--keyword-map",
        default=str(Path("ops/agent-native-revenue/seo-keyword-map.csv")),
        help="CSV source for the Keyword Map tab.",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    sources = {
        "Product Metrics": Path(args.product_metrics),
        "QA Checklist": Path(args.qa_checklist),
        "Task Queue": Path(args.task_queue),
        "Linear Import": Path(args.linear_import),
        "Keyword Map": Path(args.keyword_map),
    }

    for sheet_name, path in sources.items():
        if not path.exists():
            raise SystemExit(f"Missing CSV for {sheet_name!r}: {path}")

    workbook = load_workbook(xlsx_path)
    for sheet_name, path in sources.items():
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"Sheet not found: {sheet_name!r}. Available: {workbook.sheetnames}")
        rows = read_csv_rows(path)
        write_rows_preserve_dimensions(workbook[sheet_name], rows)
        print(f"Updated {sheet_name} from {path.as_posix()}")

    workbook.save(xlsx_path)
    print(f"Saved {xlsx_path.as_posix()}")


if __name__ == "__main__":
    main()

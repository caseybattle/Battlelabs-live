import argparse
import csv
import os
from pathlib import Path

from openpyxl import load_workbook


def cell_to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def export_sheet_to_csv(workbook_path: Path, sheet_name: str, out_path: Path) -> None:
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name!r}. Available: {workbook.sheetnames}")

    sheet = workbook[sheet_name]
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        out_path.write_text("", encoding="utf-8")
        return

    def is_empty(value: object) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    max_cols = 0
    for row in rows:
        last_non_empty = 0
        for idx, value in enumerate(row, start=1):
            if not is_empty(value):
                last_non_empty = idx
        if last_non_empty > max_cols:
            max_cols = last_non_empty

    with out_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            trimmed = list(row[:max_cols])
            padded = trimmed + [None] * (max_cols - len(trimmed))
            writer.writerow([cell_to_str(value) for value in padded])


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export battlelabs-agent-product-dashboard.xlsx tabs to CSV for easy Google Sheets import.",
    )
    parser.add_argument(
        "--xlsx",
        default=str(Path("ops/agent-native-revenue/battlelabs-agent-product-dashboard.xlsx")),
        help="Path to the dashboard XLSX.",
    )
    parser.add_argument(
        "--out-dir",
        default=str(Path("ops/agent-native-revenue/dashboard-tabs")),
        help="Directory to write exported CSV files into.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.xlsx)
    if not workbook_path.exists():
        raise SystemExit(f"XLSX not found: {workbook_path}")

    out_dir = Path(args.out_dir)

    workbook = load_workbook(workbook_path, data_only=True)
    for sheet_name in workbook.sheetnames:
        safe_name = "".join(ch if ch.isalnum() or ch in (" ", "-", "_") else "_" for ch in sheet_name).strip()
        out_path = out_dir / f"{safe_name}.csv"
        export_sheet_to_csv(workbook_path, sheet_name, out_path)
        print(f"Wrote {out_path.as_posix()} ({sheet_name})")


if __name__ == "__main__":
    main()

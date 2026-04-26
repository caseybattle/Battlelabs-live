from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph


ROOT = Path(__file__).resolve().parents[1]
PRODUCT_DIR = ROOT / "public" / "products" / "memory-journal-gift-kit"
OPS_DIR = ROOT / "ops" / "agent-native-revenue"


PROMPTS = [
    "A favorite meal we shared and why it still matters",
    "The first home you remember clearly",
    "A family saying, lesson, or tradition worth preserving",
    "A difficult season that taught you something useful",
    "A person who shaped your life in a quiet way",
    "A holiday memory that feels like home",
    "A recipe, routine, or habit you hope gets passed down",
    "A story about work, money, or making do",
    "A place you wish the family could see through your eyes",
    "Advice you would give to the next generation",
    "A photo you would choose for this page and the story behind it",
    "A moment when you felt proud of the family",
]


def ensure_dirs() -> None:
    PRODUCT_DIR.mkdir(parents=True, exist_ok=True)
    OPS_DIR.mkdir(parents=True, exist_ok=True)


def draw_wrapped_text(
    pdf: canvas.Canvas,
    text: str,
    x: float,
    y: float,
    width: float,
    style: ParagraphStyle,
) -> float:
    paragraph = Paragraph(text, style)
    _, height = paragraph.wrap(width, 8 * inch)
    paragraph.drawOn(pdf, x, y - height)
    return y - height


def draw_journal_page(pdf: canvas.Canvas, page_number: int, title: str, prompt: str) -> None:
    width, height = letter
    margin = 0.72 * inch
    pdf.setFillColor(colors.HexColor("#f8f1e7"))
    pdf.rect(0, 0, width, height, fill=1, stroke=0)
    pdf.setStrokeColor(colors.HexColor("#36584c"))
    pdf.setLineWidth(1.2)
    pdf.roundRect(margin / 2, margin / 2, width - margin, height - margin, 18, stroke=1, fill=0)

    pdf.setFillColor(colors.HexColor("#1f2a29"))
    pdf.setFont("Times-Bold", 22)
    pdf.drawString(margin, height - margin - 12, title)

    pdf.setFillColor(colors.HexColor("#a4563b"))
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(margin, height - margin - 34, f"Memory prompt {page_number:03d}")

    styles = getSampleStyleSheet()
    prompt_style = ParagraphStyle(
        "Prompt",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#253331"),
    )
    y = draw_wrapped_text(pdf, prompt, margin, height - margin - 68, width - 2 * margin, prompt_style)

    pdf.setStrokeColor(colors.HexColor("#d1b99a"))
    line_y = y - 28
    for _ in range(14):
        pdf.line(margin, line_y, width - margin, line_y)
        line_y -= 28

    pdf.setFillColor(colors.HexColor("#6b766f"))
    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(width - margin, margin - 4, "Battlelabs Memory Journal Gift Kit")
    pdf.showPage()


def create_journal_pdf(path: Path, total_pages: int, sample: bool) -> None:
    pdf = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter

    pdf.setFillColor(colors.HexColor("#243d37"))
    pdf.rect(0, 0, width, height, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#f8f1e7"))
    pdf.setFont("Times-Bold", 34)
    pdf.drawCentredString(width / 2, height - 2.2 * inch, "Grandma's Story Keeper")
    pdf.setFont("Helvetica", 13)
    pdf.drawCentredString(width / 2, height - 2.55 * inch, "A guided memory journal gift interior")
    pdf.setFillColor(colors.HexColor("#e7b06f"))
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(
        width / 2,
        height - 2.9 * inch,
        "FREE SAMPLE" if sample else "FULL PRINTABLE INTERIOR",
    )
    pdf.showPage()

    for index in range(1, total_pages):
        prompt = PROMPTS[(index - 1) % len(PROMPTS)]
        title = "Story Page" if index % 5 else "Family Keepsake Page"
        draw_journal_page(pdf, index, title, prompt)

    pdf.save()


def create_preview_pdf(path: Path) -> None:
    pdf = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=12,
        leading=18,
        textColor=colors.HexColor("#253331"),
    )
    pdf.setFillColor(colors.HexColor("#f8f1e7"))
    pdf.rect(0, 0, width, height, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#243d37"))
    pdf.setFont("Helvetica-Bold", 26)
    pdf.drawString(0.75 * inch, height - 1.1 * inch, "Memory Journal Gift Launch Kit")
    y = height - 1.55 * inch
    for text in [
        "A ready-to-package printable interior and launch kit for sellers building family memory journal products.",
        "Includes a free sample, full printable interior, keyword scorecard, title angles, listing copy, and launch checklist.",
        "Use it to validate a narrow giftable journal concept before investing in a larger KDP or printable product line.",
    ]:
        y = draw_wrapped_text(pdf, text, 0.75 * inch, y, width - 1.5 * inch, body) - 14
    pdf.setFillColor(colors.HexColor("#a4563b"))
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(0.75 * inch, y - 10, "Intro price target: $9")
    pdf.drawString(0.75 * inch, y - 34, "Regular price target: $19")
    pdf.save()


def create_csv_assets() -> None:
    with (PRODUCT_DIR / "kdp-niche-scorecard.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Niche", "Buyer Intent", "Giftability", "Competition", "Differentiation", "Build Ease", "Score"])
        writer.writerow(["Grandma memory journal", 9, 10, 6, 8, 9, 42])
        writer.writerow(["Grandpa memory journal", 8, 10, 6, 8, 9, 41])
        writer.writerow(["Mother daughter memory journal", 8, 9, 7, 7, 9, 40])
        writer.writerow(["Family recipe memory journal", 7, 9, 5, 9, 8, 38])

    with (PRODUCT_DIR / "launch-checklist.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Stage", "Task", "Owner Agent", "Status"])
        rows = [
            ("Research", "Validate narrow recipient angle", "Opportunity Agent", "Ready"),
            ("Product", "Review interior for duplicate or low-value pages", "QA Agent", "Ready"),
            ("Listing", "Select title/subtitle without income or ranking claims", "Build Agent", "Ready"),
            ("Mockups", "Export cover and preview graphics", "Design Agent", "Ready"),
            ("Launch", "Publish landing page and connect checkout", "Controller Agent", "Ready"),
            ("Measure", "Track views, opt-ins, checkout clicks, and sales", "Controller Agent", "Ready"),
        ]
        writer.writerows(rows)


def create_markdown_copy() -> None:
    copy = """# Memory Journal Gift Launch Kit

## Product promise
Package a narrow family memory journal idea faster with a printable interior, launch checklist, listing copy, and niche scorecard.

## Title angles
- Grandma's Story Keeper: A Guided Memory Journal for Family Stories
- Letters From Grandma: Prompts for Memories, Recipes, Lessons, and Love
- The Grandma Keepsake Journal: Guided Pages for the Stories We Never Want to Lose

## Short listing description
Create a thoughtful family keepsake product with a ready-to-review printable memory journal interior and launch kit. This bundle is built for creators testing narrow giftable journals before spending time on a full product line.

## What's included
- Free 20-page sample PDF
- Full 120-page printable interior PDF
- KDP niche scorecard CSV
- Title and subtitle angles
- Listing description copy
- Launch checklist
- Preview PDF and promo graphics

## Compliance notes
- Review all generated text and design before publishing.
- Disclose AI-generated content on platforms that require it.
- Do not copy trademarked phrases, brand names, or competitor listings.
- No income claims, ranking guarantees, or fake reviews.
"""
    (PRODUCT_DIR / "listing-copy.md").write_text(copy, encoding="utf-8")


def create_svg_assets() -> None:
    pins = [
        ("gift-kit-pin-01.svg", "Grandma Memory Journal Kit", "Printable interior + launch checklist"),
        ("gift-kit-pin-02.svg", "Start with one giftable niche", "Score demand before building a catalog"),
        ("gift-kit-pin-03.svg", "Family stories sell as keepsakes", "Prompts, pages, titles, mockups"),
        ("gift-kit-pin-04.svg", "KDP-ready planning assets", "No fake claims. No generic filler."),
        ("gift-kit-pin-05.svg", "Agent-built product sprint", "Build, publish, measure, clone"),
    ]
    for filename, headline, subhead in pins:
        svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="1080" height="1350" viewBox="0 0 1080 1350">
  <rect width="1080" height="1350" fill="#f8f1e7"/>
  <rect x="76" y="84" width="928" height="1182" rx="36" fill="#243d37"/>
  <circle cx="890" cy="224" r="92" fill="#e7b06f"/>
  <rect x="148" y="238" width="560" height="34" rx="17" fill="#a4563b"/>
  <text x="148" y="410" font-family="Georgia, serif" font-size="88" font-weight="700" fill="#f8f1e7">{headline}</text>
  <text x="148" y="560" font-family="Arial, sans-serif" font-size="42" fill="#d8c7ad">{subhead}</text>
  <rect x="148" y="790" width="784" height="260" rx="26" fill="#f8f1e7"/>
  <text x="194" y="895" font-family="Arial, sans-serif" font-size="44" font-weight="700" fill="#243d37">Free sample + paid kit</text>
  <text x="194" y="970" font-family="Arial, sans-serif" font-size="34" fill="#36584c">Battlelabs agent-built tools</text>
  <text x="148" y="1170" font-family="Arial, sans-serif" font-size="30" fill="#e7b06f">battlelabs.live/memory-journal-gift-kit</text>
</svg>
"""
        (PRODUCT_DIR / filename).write_text(svg, encoding="utf-8")


def create_tracking_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Metrics"
    ws.append(
        [
            "Product",
            "Views",
            "Sample Downloads",
            "Checkout Clicks",
            "Sales",
            "Refunds",
            "Gross Revenue",
            "Net Revenue",
            "Build Hours",
            "Revenue Per Build Hour",
            "Conversion Rate",
            "Next Action",
        ]
    )
    ws.append(["Memory Journal Gift Launch Kit", 0, 0, 0, 0, 0, 0, 0, 0, "=IF(I2>0,H2/I2,0)", "=IF(B2>0,E2/B2,0)", "Launch"])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="243D37")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max(14, min(max_len + 2, 28))

    qa = wb.create_sheet("QA Checklist")
    qa.append(["Check", "Status", "Notes"])
    checks = [
        "AI disclosure reviewed for target platform",
        "No income, ranking, or sales guarantee",
        "No copied competitor listing text",
        "PDF opens and page count is correct",
        "Checkout link tested before publication",
        "Refund terms visible on landing page",
    ]
    for check in checks:
        qa.append([check, "Open", ""])
    for cell in qa[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="A4563B")
    qa.freeze_panes = "A2"
    qa.column_dimensions["A"].width = 48
    qa.column_dimensions["B"].width = 14
    qa.column_dimensions["C"].width = 36

    wb.save(OPS_DIR / "battlelabs-agent-product-dashboard.xlsx")


def main() -> None:
    ensure_dirs()
    create_journal_pdf(PRODUCT_DIR / "grandma-memory-journal-free-sample.pdf", 20, sample=True)
    create_journal_pdf(PRODUCT_DIR / "grandma-memory-journal-full-interior.pdf", 120, sample=False)
    create_preview_pdf(PRODUCT_DIR / "product-preview.pdf")
    create_csv_assets()
    create_markdown_copy()
    create_svg_assets()
    create_tracking_workbook()


if __name__ == "__main__":
    main()
